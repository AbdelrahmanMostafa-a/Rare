"""
excel_export.py
بناء تقرير Excel تفاعلي (فلاتر + تجميد الصف العلوي + ألوان منظمة حسب الحالة)
من نتيجة المقارنة.
"""

from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from debt_core import (
    FOLLOWUP_SHEET_NAME,
    NOT_SENT_STATUS,
    OVERDUE_COL,
    ROUTE_COL,
    build_followup_list,
    compute_route_coverage,
)

NUMERIC_COLS = ["صافي المديونيه", "رصيد PDF", "الفرق", OVERDUE_COL]
INTEGER_COLS = ["الترتيب"]
PERCENT_COLS = ["نسبة تغطية الخط %", "نسبة التغطية %"]

STATUS_STYLES = {
    "مطابق":       {"row_fill": "C6EFCE", "header_fill": "375623", "tab_color": "70AD47"},
    "يوجد فرق":    {"row_fill": "FFC7CE", "header_fill": "C00000", "tab_color": "C00000"},
    "PDF فقط":     {"row_fill": "BDD7EE", "header_fill": "1F4E78", "tab_color": "2E75B6"},
    "مديونية فقط": {"row_fill": "FFEB9C", "header_fill": "7F6000", "tab_color": "BF8F00"},
    NOT_SENT_STATUS: {"row_fill": "D9D9D9", "header_fill": "595959", "tab_color": "A6A6A6"},
}
FOLLOWUP_STYLE = {"row_fill": "FFD966", "header_fill": "BF6000", "tab_color": "ED7D31"}
INFO_STYLE = {"header_fill": "1C1C1C", "tab_color": "808080"}
COVERAGE_STYLE = {"header_fill": "31859C", "tab_color": "31859C"}
INFO_SHEET_NAME = "معلومات التقرير"
COVERAGE_SHEET_NAME = "تغطية الخطوط"
DEFAULT_HEADER_FILL = "404040"

_thin = Side(style="thin", color="B7B7B7")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _style_header(ws, ncols, fill_hex):
    fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = BORDER


def _safe_str(v):
    return "" if pd.isna(v) else str(v)


def _autosize_columns(ws, df):
    for i, col in enumerate(df.columns, start=1):
        values = [_safe_str(v) for v in df[col].tolist()]
        max_len = max([len(str(col))] + [len(v) for v in values])
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 3, 35)


def _apply_data_styling(ws, df, row_fill_hex=None, by_status=False):
    nrows, ncols = df.shape
    numeric_positions = {
        i + 1 for i, col in enumerate(df.columns) if col in NUMERIC_COLS
    }
    integer_positions = {
        i + 1 for i, col in enumerate(df.columns) if col in INTEGER_COLS
    }
    percent_positions = {
        i + 1 for i, col in enumerate(df.columns) if col in PERCENT_COLS
    }
    status_values = df["الحالة"].tolist() if (by_status and "الحالة" in df.columns) else None

    for offset in range(nrows):
        row_num = offset + 2

        if status_values is not None:
            style = STATUS_STYLES.get(status_values[offset])
            fill_hex = style["row_fill"] if style else None
        else:
            fill_hex = row_fill_hex

        fill = (
            PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
            if fill_hex else None
        )

        for c in range(1, ncols + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.border = BORDER
            if fill:
                cell.fill = fill
            if c in numeric_positions:
                cell.number_format = "#,##0.00"
            elif c in percent_positions:
                cell.number_format = "0.0"
            elif c in integer_positions:
                cell.number_format = "#,##0"


def _finalize_sheet(ws, df, header_fill_hex, tab_color=None, row_fill_hex=None, by_status=False):
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    if tab_color:
        ws.sheet_properties.tabColor = tab_color

    _style_header(ws, df.shape[1], header_fill_hex)
    _apply_data_styling(ws, df, row_fill_hex=row_fill_hex, by_status=by_status)
    _autosize_columns(ws, df)


def _build_info_df(result, meta):
    meta = meta or {}
    counts = result["الحالة"].value_counts() if "الحالة" in result.columns else {}

    rows = [
        ("تاريخ ووقت إنشاء التقرير", meta.get("generated_at", "")),
        ("ملف المديونية المستخدم", meta.get("debt_file_name", "")),
        ("عدد عملاء ملف المديونية", meta.get("debt_rows", "")),
        ("عدد ملفات PDF المستخدمة", len(meta.get("pdf_file_names", []))),
        ("أسماء ملفات PDF", "، ".join(meta.get("pdf_file_names", []) or [])),
        ("إجمالي عدد العملاء في التقرير", len(result)),
        ("عدد العملاء مطابق", int(counts.get("مطابق", 0))),
        ("عدد العملاء يوجد فرق", int(counts.get("يوجد فرق", 0))),
        ("عدد العملاء PDF فقط", int(counts.get("PDF فقط", 0))),
        ("عدد العملاء مديونية فقط", int(counts.get("مديونية فقط", 0))),
        ("عدد العملاء (الخط لم يُرسل)", int(counts.get(NOT_SENT_STATUS, 0))),
    ]
    return pd.DataFrame(rows, columns=["البيان", "القيمة"])


def build_excel_report(result: pd.DataFrame, meta: dict = None) -> bytes:
    """يبني ملف Excel فيه شيت 'معلومات التقرير'، شيت 'الكل' مع فلاتر وألوان
    لكل حالة، شيت 'عملاء يجب المرور عليهم' (لو عمود التجاوز متاح)، شيت
    'تغطية الخطوط' (لو عمود الخط متاح)، وشيتات منفصلة لكل حالة."""

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        info_df = _build_info_df(result, meta)
        info_df.to_excel(writer, sheet_name=INFO_SHEET_NAME, index=False)

        result.to_excel(writer, sheet_name="الكل", index=False)

        followup_df = build_followup_list(result)
        if followup_df is not None:
            followup_df.to_excel(writer, sheet_name=FOLLOWUP_SHEET_NAME, index=False)

        coverage_df = compute_route_coverage(result)
        if coverage_df is not None:
            coverage_df.to_excel(writer, sheet_name=COVERAGE_SHEET_NAME, index=False)

        statuses = ["مطابق", "يوجد فرق", "PDF فقط", "مديونية فقط"]
        if NOT_SENT_STATUS in result["الحالة"].unique():
            statuses.append(NOT_SENT_STATUS)

        category_frames = {}
        for status in statuses:
            subset = result[result["الحالة"] == status].copy()
            # "مديونية فقط" مرتبة من الخط الأعلى تغطية للأقل، عشان العملاء
            # اللي خطهم فعلاً اتغطى (وهما غايبين) يظهروا الأول
            if status == "مديونية فقط" and "نسبة تغطية الخط %" in subset.columns:
                subset = subset.sort_values("نسبة تغطية الخط %", ascending=False)
            subset = subset.reset_index(drop=True)
            subset.to_excel(writer, sheet_name=status, index=False)
            category_frames[status] = subset

        wb = writer.book

        _finalize_sheet(
            wb[INFO_SHEET_NAME], info_df,
            header_fill_hex=INFO_STYLE["header_fill"],
            tab_color=INFO_STYLE["tab_color"],
        )

        _finalize_sheet(
            wb["الكل"], result,
            header_fill_hex=DEFAULT_HEADER_FILL,
            by_status=True,
        )

        if followup_df is not None:
            _finalize_sheet(
                wb[FOLLOWUP_SHEET_NAME], followup_df,
                header_fill_hex=FOLLOWUP_STYLE["header_fill"],
                tab_color=FOLLOWUP_STYLE["tab_color"],
                row_fill_hex=FOLLOWUP_STYLE["row_fill"],
            )

        if coverage_df is not None:
            _finalize_sheet(
                wb[COVERAGE_SHEET_NAME], coverage_df,
                header_fill_hex=COVERAGE_STYLE["header_fill"],
                tab_color=COVERAGE_STYLE["tab_color"],
            )

        for status, subset in category_frames.items():
            style = STATUS_STYLES[status]
            _finalize_sheet(
                wb[status], subset,
                header_fill_hex=style["header_fill"],
                tab_color=style["tab_color"],
                row_fill_hex=style["row_fill"],
            )

        wb.active = wb.sheetnames.index(INFO_SHEET_NAME)

    return output.getvalue()
