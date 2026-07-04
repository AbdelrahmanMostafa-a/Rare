"""
excel_export.py
بناء تقرير Excel تفاعلي (فلاتر + تجميد الصف العلوي + ألوان منظمة حسب الحالة)
من نتيجة المقارنة.
"""

from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NUMERIC_COLS = ["صافي المديونيه", "رصيد PDF", "الفرق"]

STATUS_STYLES = {
    "مطابق":       {"row_fill": "C6EFCE", "header_fill": "375623", "tab_color": "70AD47"},
    "يوجد فرق":    {"row_fill": "FFC7CE", "header_fill": "C00000", "tab_color": "C00000"},
    "PDF فقط":     {"row_fill": "BDD7EE", "header_fill": "1F4E78", "tab_color": "2E75B6"},
    "مديونية فقط": {"row_fill": "FFEB9C", "header_fill": "7F6000", "tab_color": "BF8F00"},
}
DEFAULT_HEADER_FILL = "404040"

_thin = Side(style="thin", color="B7B7B7")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _style_header(ws, ncols, fill_hex):
    fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = BORDER


def _safe_str(v):
    return "" if pd.isna(v) else str(v)


def _autosize_columns(ws, df):
    for i, col in enumerate(df.columns, start=1):
        values = [_safe_str(v) for v in df[col].tolist()]
        max_len = max([len(str(col))] + [len(v) for v in values])
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 3, 35)


def _apply_data_styling(ws, df, row_fill_hex=None, by_status=False):
    nrows, ncols = df.shape
    numeric_positions = {
        i + 1 for i, col in enumerate(df.columns) if col in NUMERIC_COLS
    }
    status_values = df["الحالة"].tolist() if (by_status and "الحالة" in df.columns) else None

    for offset in range(nrows):
        row_num = offset + 2

        if status_values is not None:
            style = STATUS_STYLES.get(status_values[offset])
            fill_hex = style["row_fill"] if style else None
        else:
            fill_hex = row_fill_hex

        fill = (
            PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
            if fill_hex else None
        )

        for c in range(1, ncols + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.border = BORDER
            if fill:
                cell.fill = fill
            if c in numeric_positions:
                cell.number_format = "#,##0.00"


def _finalize_sheet(ws, df, header_fill_hex, tab_color=None, row_fill_hex=None, by_status=False):
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    if tab_color:
        ws.sheet_properties.tabColor = tab_color

    _style_header(ws, df.shape[1], header_fill_hex)
    _apply_data_styling(ws, df, row_fill_hex=row_fill_hex, by_status=by_status)
    _autosize_columns(ws, df)


def build_excel_report(result: pd.DataFrame) -> bytes:
    """يبني ملف Excel فيه شيت 'الكل' مع فلاتر وألوان لكل حالة، وشيتات منفصلة لكل حالة."""

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        result.to_excel(writer, sheet_name="الكل", index=False)

        category_frames = {}
        for status in ["مطابق", "يوجد فرق", "PDF فقط", "مديونية فقط"]:
            subset = result[result["الحالة"] == status].reset_index(drop=True)
            subset.to_excel(writer, sheet_name=status, index=False)
            category_frames[status] = subset

        wb = writer.book

        _finalize_sheet(
            wb["الكل"], result,
            header_fill_hex=DEFAULT_HEADER_FILL,
            by_status=True,
        )

        for status, subset in category_frames.items():
            style = STATUS_STYLES[status]
            _finalize_sheet(
                wb[status], subset,
                header_fill_hex=style["header_fill"],
                tab_color=style["tab_color"],
                row_fill_hex=style["row_fill"],
            )

    return output.getvalue()
