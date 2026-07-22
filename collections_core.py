# -*- coding: utf-8 -*-
"""
collections_core.py
أداة منفصلة تمامًا عن أداة مقارنة المديونية: بتقرأ ملفات PDF من نوع
"سجل المعاملات" / "الحساب ديالك" (كشوف تحصيل يومية لمندوب أو حساب معين)،
وتفصل اسم العميل عن رقمه، وتجمع إجمالي التحصيل لكل عميل، وإجمالي كل ملف.
"""

import re
import unicodedata
from io import BytesIO

import pandas as pd
import pdfplumber
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DATE_RE = re.compile(r"(\d{4}-\d{2}-\d{2})")
NUM_RE = re.compile(r"(470\d{7,8})")
AMOUNT_RE = re.compile(r"^\d+,\d{2}$")
ARABIC_CHARS_RE = re.compile(r"[\u0600-\u06FF\uFB50-\uFDFF\uFE70-\uFEFF]")


def _fix_mixed_ascii(s):
    # بعد عكس السطر، أي كلمة إنجليزية جواه بتترجع تتقلب تاني بالغلط
    # (لأنها كانت أصلاً مكتوبة صح) - نرجعها لوضعها الطبيعي
    return re.sub(r"[A-Za-z]+", lambda m: m.group(0)[::-1], s)


def _smart_readable(s):
    """يعكس النص بس لو فيه حروف عربية (عشان متبوظش أرقام أو نصوص إنجليزية خالصة)."""
    s = s.strip()
    if not s:
        return s
    if ARABIC_CHARS_RE.search(s):
        return _fix_mixed_ascii(unicodedata.normalize("NFKC", s[::-1]))
    return s


def extract_collector_label(file):
    """بياخد اسم/رقم صاحب الكشف والفترة الزمنية من أول الصفحة الأولى، لعرضها كعنوان."""
    with pdfplumber.open(file) as pdf:
        text = pdf.pages[0].extract_text() or ""

    lines = [l.strip() for l in text.split("\n") if l.strip()]

    name_parts = []
    for line in lines[:4]:
        if DATE_RE.search(line) or "الحساب" in _smart_readable(line) or "سجل" in _smart_readable(line):
            break
        name_parts.append(_smart_readable(line))

    label = " / ".join(dict.fromkeys([p for p in name_parts if p]))

    date_range = ""
    m = re.search(r"(\d{2}-\d{2}-\d{4}).{0,10}?(\d{2}-\d{2}-\d{4})", text)
    if m:
        date_range = f"{m.group(1)} → {m.group(2)}"

    if label and date_range:
        return f"{label} ({date_range})"
    return label or date_range or "ملف بدون عنوان"


def parse_transaction_pdf(file):
    """
    يقرأ ملف PDF من نوع سجل المعاملات/الحساب ديالك ويرجع DataFrame بكل
    معاملة على حدة: التاريخ، رقم العميل، اسم العميل، تحصيل (أخذت)، معطى (أعطيت).
    """
    rows = []
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.split("\n"):
                date_m = DATE_RE.search(line)
                num_m = NUM_RE.search(line)
                if not date_m or not num_m:
                    continue

                customer_number = num_m.group(1)
                before_date = line[:date_m.start()].strip()
                tokens = before_date.split()
                if len(tokens) < 3:
                    continue

                amt1, amt2 = tokens[0], tokens[1]
                taken = float(amt1.replace(",", ".")) if AMOUNT_RE.match(amt1) else 0.0
                given = float(amt2.replace(",", ".")) if AMOUNT_RE.match(amt2) else 0.0

                rest = " ".join(tokens[2:]).replace(customer_number, "").strip()
                name = _smart_readable(rest)

                rows.append({
                    "التاريخ": date_m.group(1),
                    "رقم العميل": customer_number,
                    "اسم العميل": name,
                    "تحصيل (أخذت)": taken,
                    "معطى (أعطيت)": given,
                })

    return pd.DataFrame(rows, columns=["التاريخ", "رقم العميل", "اسم العميل", "تحصيل (أخذت)", "معطى (أعطيت)"])


def summarize_by_customer(transactions):
    """يجمع المعاملات لكل عميل: عدد المعاملات، إجمالي التحصيل، إجمالي المعطى - من الأكبر تحصيلًا للأصغر."""
    if transactions.empty:
        return transactions.assign(**{"عدد المعاملات": []})

    names = transactions.groupby("رقم العميل")["اسم العميل"].agg(
        lambda s: next((v for v in s if v), "")
    )
    agg = transactions.groupby("رقم العميل").agg(
        **{
            "عدد المعاملات": ("رقم العميل", "count"),
            "إجمالي التحصيل": ("تحصيل (أخذت)", "sum"),
            "إجمالي المعطى": ("معطى (أعطيت)", "sum"),
        }
    )
    agg["اسم العميل"] = names
    agg = agg.reset_index()[
        ["رقم العميل", "اسم العميل", "عدد المعاملات", "إجمالي التحصيل", "إجمالي المعطى"]
    ]
    return agg.sort_values("إجمالي التحصيل", ascending=False).reset_index(drop=True)


def combine_all_summaries(per_pdf_results):
    """يدمج جداول تلخيص كل الملفات في جدول واحد مجمّع لكل عميل (لو نفس العميل
    ظهر في أكتر من ملف، بيتجمع إجماليه)، مرتب من الأكبر تحصيلًا للأصغر."""
    frames = [item["summary"] for item in per_pdf_results if not item["summary"].empty]

    cols = ["رقم العميل", "اسم العميل", "عدد المعاملات", "إجمالي التحصيل", "إجمالي المعطى"]
    if not frames:
        return pd.DataFrame(columns=cols)

    combined = pd.concat(frames, ignore_index=True)
    names = combined.groupby("رقم العميل")["اسم العميل"].agg(
        lambda s: next((v for v in s if v), "")
    )
    agg = combined.groupby("رقم العميل").agg(
        **{
            "عدد المعاملات": ("عدد المعاملات", "sum"),
            "إجمالي التحصيل": ("إجمالي التحصيل", "sum"),
            "إجمالي المعطى": ("إجمالي المعطى", "sum"),
        }
    )
    agg["اسم العميل"] = names
    agg = agg.reset_index()[cols]
    return agg.sort_values("إجمالي التحصيل", ascending=False).reset_index(drop=True)


# ---------------------------------------------------------------------------
# تصدير Excel
# ---------------------------------------------------------------------------
_thin = Side(style="thin", color="B7B7B7")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
HEADER_FILL = "1F4E78"
TOTAL_FILL = "D9E1F2"


def _style_sheet(ws, ncols, nrows):
    header_fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")

    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align
        cell.border = BORDER

    for r in range(2, nrows + 2):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).border = BORDER

    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_block(ws, start_row, label, df):
    """يكتب عنوان + جدول + صف إجمالي لملف واحد، ويرجع رقم الصف بعد الفاصل."""
    title_font = Font(bold=True, size=13, color="1F4E78")
    header_fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    total_fill = PatternFill(start_color=TOTAL_FILL, end_color=TOTAL_FILL, fill_type="solid")
    money_cols = {"إجمالي التحصيل", "إجمالي المعطى"}

    row = start_row
    ws.cell(row=row, column=1, value=f"📄 {label}").font = title_font
    row += 1

    if df.empty:
        ws.cell(row=row, column=1, value="لا توجد بيانات في الملف ده")
        return row + 2

    for c, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=row, column=c, value=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")
    row += 1

    for _, data_row in df.iterrows():
        for c, col in enumerate(df.columns, start=1):
            val = data_row[col]
            cell = ws.cell(row=row, column=c, value=(None if pd.isna(val) else val))
            cell.border = BORDER
            if col in money_cols:
                cell.number_format = "#,##0.00"
        row += 1

    col_idx = {col: i + 1 for i, col in enumerate(df.columns)}
    ws.cell(row=row, column=1, value="الإجمالي")
    if "عدد المعاملات" in col_idx:
        ws.cell(row=row, column=col_idx["عدد المعاملات"], value=int(df["عدد المعاملات"].sum()))
    for money_col in money_cols:
        if money_col in col_idx:
            cell = ws.cell(row=row, column=col_idx[money_col], value=float(df[money_col].sum()))
            cell.number_format = "#,##0.00"
    for c in range(1, len(df.columns) + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.border = BORDER

    return row + 2  # سطر فاضي فاصل بين الملفات


def build_collections_excel(per_pdf_results):
    """
    يبني ملف Excel من شيتين بس:
    - شيت "تفصيل كل ملف": كل ملف وجدوله وإجماليه لوحدهم، تحت بعض في نفس الشيت
    - شيت "الإجمالي المجمّع": كل العملاء من كل الملفات مجمّعين في جدول واحد
    """
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        wb = writer.book

        # ---- شيت 1: كل تحليل لوحده تحت بعض ----
        ws1 = wb.create_sheet("تفصيل كل ملف")
        ws1.sheet_view.rightToLeft = True
        writer.sheets[ws1.title] = ws1

        row = 1
        for item in per_pdf_results:
            row = _write_block(ws1, row, item["label"], item["summary"])

        max_cols = max(
            (len(item["summary"].columns) for item in per_pdf_results if not item["summary"].empty),
            default=5,
        )
        for c in range(1, max_cols + 1):
            ws1.column_dimensions[get_column_letter(c)].width = 24

        # ---- شيت 2: كل التحليلات مجمعة في جدول واحد ----
        combined = combine_all_summaries(per_pdf_results)
        sheet2_name = "الإجمالي المجمع"
        combined.to_excel(writer, sheet_name=sheet2_name, index=False)
        ws2 = writer.sheets[sheet2_name]
        nrows, ncols = combined.shape

        if nrows > 0:
            total_row = nrows + 2
            col_idx2 = {col: i + 1 for i, col in enumerate(combined.columns)}
            ws2.cell(row=total_row, column=1, value="الإجمالي")
            if "عدد المعاملات" in col_idx2:
                ws2.cell(row=total_row, column=col_idx2["عدد المعاملات"], value=int(combined["عدد المعاملات"].sum()))
            for money_col in ("إجمالي التحصيل", "إجمالي المعطى"):
                if money_col in col_idx2:
                    cell = ws2.cell(row=total_row, column=col_idx2[money_col], value=float(combined[money_col].sum()))
                    cell.number_format = "#,##0.00"

            total_fill = PatternFill(start_color=TOTAL_FILL, end_color=TOTAL_FILL, fill_type="solid")
            for c in range(1, ncols + 1):
                cell = ws2.cell(row=total_row, column=c)
                cell.font = Font(bold=True)
                cell.fill = total_fill
                cell.border = BORDER

            money_positions = {
                i + 1 for i, col in enumerate(combined.columns)
                if col in ("إجمالي التحصيل", "إجمالي المعطى")
            }
            for c in money_positions:
                for r in range(2, total_row + 1):
                    ws2.cell(row=r, column=c).number_format = "#,##0.00"

            _style_sheet(ws2, ncols, nrows)

        for i, col in enumerate(combined.columns, start=1):
            values = [("" if pd.isna(v) else str(v)) for v in combined[col].tolist()]
            max_len = max([len(str(col))] + [len(v) for v in values]) if values else len(str(col))
            ws2.column_dimensions[get_column_letter(i)].width = min(max_len + 3, 35)

    return output.getvalue()
